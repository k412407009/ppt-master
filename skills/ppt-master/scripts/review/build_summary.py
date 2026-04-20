"""
Review Board · 跨项目评审汇总 (review-summary.md + projects_comparison.xlsx)

读 <projects_root>/<project_dir>/review/*_review.json 全部
产出:
  - <projects_root>/review-summary.md              人看的 markdown 汇总
  - <projects_root>/projects_comparison.xlsx       机器看/可筛选的 4-sheet 对比 Excel

Usage:
    python skills/ppt-master/scripts/review/build_summary.py <projects_root>
"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
except Exception:
    pass

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


DIMENSIONS = {
    "D1": "战略-题材匹配度",
    "D2": "玩法-核心循环",
    "D3": "玩法-时间节点",
    "D4": "玩法-阶段过渡",
    "D5": "商业化-付费/留存",
    "D6": "风险-题材/合规",
    "D7": "美术/配色/素材",
    "D8": "落地-团队/排期/预算",
    "D9": "演讲-PPT 表达力",
}

VERDICT_LABEL = {
    "pass": "✅ PASS",
    "conditional_pass": "⚠️ CONDITIONAL PASS",
    "not_pass": "❌ REJECT",
}

VERDICT_RANK = {"pass": 0, "conditional_pass": 1, "not_pass": 2}

# Run 平台官方排级 (外部题材数据, 定期跟用户同步; 找不到就显示 '-')
# 优先级: review.json 里的 run_platform_rank 字段 > 这里的 fallback 映射
RUN_RANK_FALLBACK = {
    "A": "S+ 最优",
    "G": "S 推荐",
    "E": "A+ 推荐",
    "H": "A+ 推荐",
    "I": "A 推荐",
    "D": "D 致命(买量封杀)",
}

PRIORITY_COLOR = {
    "P0": "FFCCCC",
    "P1": "FFE5B4",
    "P2": "DCE6F1",
}

VERDICT_COLOR = {
    "pass": "C6EFCE",
    "conditional_pass": "FFEB9C",
    "not_pass": "FFC7CE",
}

HEADER_FILL = PatternFill("solid", fgColor="1A2C5C")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Microsoft YaHei")
CELL_FONT = Font(name="Microsoft YaHei", size=10)
THICK_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# JSON 内部保留 P/S1/D8 等代号 (机器可读外键), 但所有面向人的产物一律展开成全名。

def _rev_label(rev: dict) -> str:
    """评委显示标签: '资深制作人 (15年)'。"""
    return f"{rev['name']} ({rev['years']}年)"


def _rev_lookup(reviewers: list[dict]) -> dict[str, dict]:
    return {r["id"]: r for r in reviewers}


def _avg_per_dim(scores: dict[str, dict[str, int]]) -> dict[str, float]:
    out = {}
    for d in DIMENSIONS:
        vals = [s.get(d, 0) for s in scores.values()]
        vals = [v for v in vals if v]
        out[d] = round(sum(vals) / len(vals), 2) if vals else 0.0
    return out


def _project_short(data: dict) -> str:
    """短名: 'A_港口开箱' (截第一个空格 / '(' 前), 用于 Excel 列头和 overview 第一列。"""
    name = data["project"]
    for sep in [" (", "(", " - ", " | "]:
        if sep in name:
            return name.split(sep)[0].strip()
    return name[:24]


def _run_rank(data: dict, project_dir_name: str) -> str:
    """读 review.json 里的 run_platform_rank; 否则按项目目录名首字母查 fallback。"""
    r = data.get("run_platform_rank")
    if r:
        return r
    first = project_dir_name[:1].upper()
    return RUN_RANK_FALLBACK.get(first, "-")


def _count_by_priority(issues: list[dict]) -> dict[str, int]:
    out = {"P0": 0, "P1": 0, "P2": 0}
    for q in issues:
        p = q.get("priority")
        if p in out:
            out[p] += 1
    return out


def _style_header_row(ws, ncols: int, row: int = 1) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THICK_BORDER


def _build_comparison_xlsx(all_data: list[dict], out_path: Path) -> None:
    """产出 4 sheet 的跨项目对比 Excel。"""
    wb = Workbook()

    sorted_data = sorted(
        all_data,
        key=lambda x: (VERDICT_RANK.get(x["verdict"], 9), -float(x["weighted_score"])),
    )

    # ============ Sheet 1: Overview ============
    ws1 = wb.active
    ws1.title = "Overview"
    h1 = [
        "项目", "加权总分", "裁决", "Run 排级",
        "P0 数", "P1 数", "P2 数", "评审日期", "下次复审", "项目目录",
    ]
    for c, h in enumerate(h1, start=1):
        ws1.cell(row=1, column=c, value=h)
    _style_header_row(ws1, len(h1))

    for r, d in enumerate(sorted_data, start=2):
        pc = _count_by_priority(d["issues"])
        row_vals = [
            _project_short(d),
            float(d["weighted_score"]),
            VERDICT_LABEL.get(d["verdict"], d["verdict"]),
            _run_rank(d, d["__project_dir"]),
            pc["P0"], pc["P1"], pc["P2"],
            d.get("review_date", "-"),
            d.get("next_review", "-"),
            d["__project_dir"],
        ]
        for c, v in enumerate(row_vals, start=1):
            cell = ws1.cell(row=r, column=c, value=v)
            cell.font = CELL_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border = THICK_BORDER
        # 裁决颜色
        ws1.cell(row=r, column=3).fill = PatternFill(
            "solid", fgColor=VERDICT_COLOR.get(d["verdict"], "FFFFFF")
        )
        # P0 数高亮红色 (≥5 表示重点项目)
        p0_cell = ws1.cell(row=r, column=5)
        if pc["P0"] >= 5:
            p0_cell.fill = PatternFill("solid", fgColor="FFCCCC")
            p0_cell.font = Font(name="Microsoft YaHei", size=10, bold=True)

    widths1 = [22, 10, 24, 22, 8, 8, 8, 14, 38, 48]
    for i, w in enumerate(widths1, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = "A2"

    # ============ Sheet 2: Scores_Matrix ============
    ws2 = wb.create_sheet("Scores_Matrix")
    short_names = [_project_short(d) for d in sorted_data]
    h2 = ["维度"] + short_names + ["跨项目均分"]
    for c, h in enumerate(h2, start=1):
        ws2.cell(row=1, column=c, value=h)
    _style_header_row(ws2, len(h2))

    avg_by_proj = {d["project"]: _avg_per_dim(d["scores"]) for d in sorted_data}
    for r, (dim_id, dim_name) in enumerate(DIMENSIONS.items(), start=2):
        ws2.cell(row=r, column=1, value=dim_name).font = Font(bold=True, name="Microsoft YaHei")
        ws2.cell(row=r, column=1).border = THICK_BORDER
        row_vals = []
        for d in sorted_data:
            v = avg_by_proj[d["project"]][dim_id]
            row_vals.append(v)
        cross_avg = round(sum(row_vals) / len(row_vals), 2) if row_vals else 0
        for c, v in enumerate(row_vals, start=2):
            cell = ws2.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(horizontal="center")
            cell.font = CELL_FONT
            cell.border = THICK_BORDER
            if v <= 2.5:
                cell.fill = PatternFill("solid", fgColor="FFC7CE")
            elif v <= 3.0:
                cell.fill = PatternFill("solid", fgColor="FFEB9C")
            elif v >= 4.5:
                cell.fill = PatternFill("solid", fgColor="C6EFCE")
        cell = ws2.cell(row=r, column=len(h2), value=cross_avg)
        cell.font = Font(bold=True, name="Microsoft YaHei")
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill("solid", fgColor="FFF7E6")
        cell.border = THICK_BORDER

    total_r = 2 + len(DIMENSIONS)
    cell = ws2.cell(row=total_r, column=1, value="加权总分")
    cell.font = Font(bold=True, name="Microsoft YaHei")
    cell.fill = PatternFill("solid", fgColor="FFD700")
    cell.border = THICK_BORDER
    for c, d in enumerate(sorted_data, start=2):
        cell = ws2.cell(row=total_r, column=c, value=float(d["weighted_score"]))
        cell.font = Font(bold=True, name="Microsoft YaHei")
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill("solid", fgColor="FFD700")
        cell.border = THICK_BORDER

    ws2.column_dimensions["A"].width = 24
    for c in range(2, len(h2) + 1):
        ws2.column_dimensions[get_column_letter(c)].width = 16
    ws2.freeze_panes = "B2"

    # ============ Sheet 3: All_Issues ============
    ws3 = wb.create_sheet("All_Issues")
    h3 = [
        "项目", "ID", "评委", "维度", "类型", "优先级", "页号", "问题", "改动建议/最优解",
    ]
    for c, h in enumerate(h3, start=1):
        ws3.cell(row=1, column=c, value=h)
    _style_header_row(ws3, len(h3))

    rownum = 2
    for d in sorted_data:
        rev_lookup = {r["id"]: r for r in d.get("reviewers", [])}
        for q in d["issues"]:
            rev = rev_lookup.get(q["reviewer"], {})
            rev_str = f"{rev.get('name', q['reviewer'])} ({rev.get('years', '-')}年)" if rev else q["reviewer"]
            advice = q.get("suggestion") or q.get("best_answer") or "-"
            row_vals = [
                _project_short(d),
                q["id"],
                rev_str,
                DIMENSIONS.get(q["dimension"], q["dimension"]),
                "客观" if q["type"] == "O" else "主观",
                q["priority"],
                q.get("page", "-"),
                q["question"],
                advice,
            ]
            for c, v in enumerate(row_vals, start=1):
                cell = ws3.cell(row=rownum, column=c, value=v)
                cell.font = CELL_FONT
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = THICK_BORDER
            ws3.cell(row=rownum, column=6).fill = PatternFill(
                "solid", fgColor=PRIORITY_COLOR.get(q["priority"], "FFFFFF")
            )
            rownum += 1

    ws3.auto_filter.ref = f"A1:{get_column_letter(len(h3))}{rownum - 1}"
    widths3 = [22, 8, 22, 22, 6, 8, 16, 50, 60]
    for i, w in enumerate(widths3, start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A2"

    # ============ Sheet 4: P0_Cross_Project ============
    ws4 = wb.create_sheet("P0_Cross_Project")
    h4 = [
        "项目", "ID", "评委", "维度", "页号", "问题", "改动建议/最优解",
    ]
    for c, h in enumerate(h4, start=1):
        ws4.cell(row=1, column=c, value=h)
    _style_header_row(ws4, len(h4))

    rownum = 2
    for d in sorted_data:
        rev_lookup = {r["id"]: r for r in d.get("reviewers", [])}
        p0_issues = [q for q in d["issues"] if q["priority"] == "P0"]
        for q in p0_issues:
            rev = rev_lookup.get(q["reviewer"], {})
            rev_str = f"{rev.get('name', q['reviewer'])} ({rev.get('years', '-')}年)" if rev else q["reviewer"]
            advice = q.get("suggestion") or q.get("best_answer") or "-"
            row_vals = [
                _project_short(d),
                q["id"],
                rev_str,
                DIMENSIONS.get(q["dimension"], q["dimension"]),
                q.get("page", "-"),
                q["question"],
                advice,
            ]
            for c, v in enumerate(row_vals, start=1):
                cell = ws4.cell(row=rownum, column=c, value=v)
                cell.font = CELL_FONT
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = THICK_BORDER
            ws4.cell(row=rownum, column=1).fill = PatternFill(
                "solid", fgColor="FFCCCC"
            )
            rownum += 1

    if rownum > 2:
        ws4.auto_filter.ref = f"A1:{get_column_letter(len(h4))}{rownum - 1}"
    widths4 = [22, 8, 22, 22, 16, 55, 65]
    for i, w in enumerate(widths4, start=1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    ws4.freeze_panes = "A2"

    wb.save(out_path)


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print("用法: python build_summary.py <projects_root>")
        return 2
    root = Path(argv[1]).resolve()
    if not root.exists():
        print(f"项目根目录不存在: {root}")
        return 2

    all_data: list[dict[str, Any]] = []
    for proj_dir in sorted(root.iterdir()):
        if not proj_dir.is_dir():
            continue
        review_dir = proj_dir / "review"
        if not review_dir.exists():
            continue
        for j in sorted(review_dir.glob("*_review.json")):
            try:
                d = json.loads(j.read_text(encoding="utf-8"))
                d["__source"] = j
                d["__project_dir"] = proj_dir.name
                all_data.append(d)
                break
            except Exception as e:
                print(f"  警告: {j} 读取失败 → {e}")

    if not all_data:
        print(f"  未找到任何 review.json (扫描: {root})")
        return 2

    print(f"  读到 {len(all_data)} 份 review.json:")
    for d in all_data:
        print(f"    - {d['project']} → {d['__source'].name}")

    # 取第一份 review 里的 reviewers 作为评委会展示 (5 人在所有项目里固定相同)
    reviewers_for_header = all_data[0].get("reviewers", [])
    rev_header = " / ".join(_rev_label(r) for r in reviewers_for_header) if reviewers_for_header else "-"

    lines: list[str] = []
    lines.append("# Review Board · 跨项目评审汇总")
    lines.append("")
    lines.append(f"> 涵盖 {len(all_data)} 个项目, 评委会 5 人: {rev_header}")
    lines.append("")
    lines.append("---")
    lines.append("")

    # ===== I. 总体裁决对照 =====
    lines.append("## I. 总体裁决对照")
    lines.append("")
    lines.append("| 项目 | 加权总分 | 裁决 | 复审节点 |")
    lines.append("|---|---|---|---|")
    for d in sorted(all_data, key=lambda x: VERDICT_RANK.get(x["verdict"], 9)):
        lines.append(
            f"| **{d['project']}** | {d['weighted_score']}/5 | "
            f"{VERDICT_LABEL.get(d['verdict'], d['verdict'])} | "
            f"{d.get('next_review', '-')} |"
        )
    lines.append("")

    # ===== II. 9 维度评分纵向对比 =====
    lines.append("## II. 9 维度评分纵向对比")
    lines.append("")
    header = ["维度"] + [d["project"][:18] for d in all_data]
    lines.append("| " + " | ".join(header) + " |")
    lines.append("|" + "|".join(["---"] * len(header)) + "|")
    avg_by_proj = {d["project"]: _avg_per_dim(d["scores"]) for d in all_data}
    for dim_id, dim_name in DIMENSIONS.items():
        row = [f"**{dim_name}**"]
        for d in all_data:
            v = avg_by_proj[d["project"]][dim_id]
            mark = ""
            if v <= 3.0:
                mark = " ⚠️"
            elif v >= 4.5:
                mark = " ⭐"
            row.append(f"{v}{mark}")
        lines.append("| " + " | ".join(row) + " |")
    overall_row = ["**加权总分**"] + [f"{d['weighted_score']}" for d in all_data]
    lines.append("| " + " | ".join(overall_row) + " |")
    lines.append("")
    lines.append("> ⚠️ = 该维度均分 ≤ 3.0 需重点回应   ⭐ = 该维度均分 ≥ 4.5 项目亮点")
    lines.append("")

    # ===== III. 共性问题 (跨 ≥2 项目) =====
    lines.append("## III. 共性问题 (≥2 项目都中招, 提到 skill 层迭代)")
    lines.append("")
    dim_lows: dict[str, list[str]] = {}
    for d in all_data:
        dims = avg_by_proj[d["project"]]
        for dim_id, v in dims.items():
            if v <= 3.5:
                dim_lows.setdefault(dim_id, []).append(d["project"])
    common_dims = {k: v for k, v in dim_lows.items() if len(v) >= 2}
    if not common_dims:
        lines.append("(无共性低分维度, 3 项目各有侧重)")
    else:
        lines.append("| 维度 | 命中项目 | 建议 |")
        lines.append("|---|---|---|")
        for dim_id, projs in common_dims.items():
            lines.append(
                f"| **{DIMENSIONS[dim_id]}** | {' / '.join(projs)} | "
                f"建议 ppt-master skill 在 strategist.md 增加该维度专项 checklist |"
            )
    lines.append("")

    # ===== IV. 各项目 P0 清单 =====
    lines.append("## IV. 各项目 P0 必改清单")
    lines.append("")
    has_p0 = False
    for d in all_data:
        p0_issues = [q for q in d["issues"] if q["priority"] == "P0"]
        if not p0_issues:
            lines.append(f"### {d['project']}")
            lines.append("")
            lines.append("✅ 无 P0 项, 可直接进入复审")
            lines.append("")
            continue
        has_p0 = True
        rev_lookup = _rev_lookup(d.get("reviewers", []))
        lines.append(f"### {d['project']} ({len(p0_issues)} 项 P0)")
        lines.append("")
        lines.append("| ID | 评委 | 影响页 | 问题 | 改动建议/最优解 |")
        lines.append("|---|---|---|---|---|")
        for q in p0_issues:
            advice = q.get("suggestion") or q.get("best_answer") or "-"
            advice = advice.replace("\n", " ").replace("|", "\\|")
            question = q["question"].replace("\n", " ").replace("|", "\\|")
            rev = rev_lookup.get(q["reviewer"], {})
            rev_str = _rev_label(rev) if rev else q["reviewer"]
            lines.append(
                f"| {q['id']} | {rev_str} | {q.get('page', '-')} | {question} | {advice} |"
            )
        lines.append("")

    # ===== V. 推进建议 =====
    lines.append("## V. 推进建议")
    lines.append("")
    rank_sorted = sorted(all_data, key=lambda x: (VERDICT_RANK.get(x["verdict"], 9), -float(x["weighted_score"])))
    lines.append("**推荐推进顺序** (从 PPT 最成熟 → 最需返工):")
    lines.append("")
    for i, d in enumerate(rank_sorted, start=1):
        p0_n = len([q for q in d["issues"] if q["priority"] == "P0"])
        lines.append(
            f"{i}. **{d['project']}** — 总分 {d['weighted_score']}/5 / "
            f"裁决 {VERDICT_LABEL.get(d['verdict'], d['verdict'])} / "
            f"P0 共 {p0_n} 项"
        )
    lines.append("")

    # ===== VI. skill 层改进建议 =====
    lines.append("## VI. ppt-master skill 层改进建议")
    lines.append("")
    if common_dims:
        lines.append("基于本轮评审, 建议 skill 在以下方面强化 (减少未来项目同样跌跟头):")
        lines.append("")
        for dim_id, projs in common_dims.items():
            lines.append(
                f"- **{DIMENSIONS[dim_id]}** → "
                f"在 `references/strategist.md` 八问八答里加一项专项校验, 命中项目: {' / '.join(projs)}"
            )
        lines.append("")
    lines.append("- 在 `references/review-board.md` §VIII 加 'verdict 提升路径' 表格 (从 conditional → pass 需要哪些动作)")
    lines.append("- 在 strategist 自查 checklist 里新增 '评委 5 视角扫雷' 条目, 让 design_spec.md 提交前自查")
    lines.append("")

    out_path = root / "review-summary.md"
    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"  wrote: {out_path}")

    xlsx_path = root / "projects_comparison.xlsx"
    _build_comparison_xlsx(all_data, xlsx_path)
    print(f"  wrote: {xlsx_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
